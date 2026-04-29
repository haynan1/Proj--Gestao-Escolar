import re
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DIAS = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']

NAVY = '111827'
NAVY_2 = '1E293B'
GREEN = '22C55E'
INK = '0F172A'
MUTED = '64748B'
LINE = 'CBD5E1'
PAGE = 'F8FAFC'
EMPTY = 'F1F5F9'
WHITE = 'FFFFFF'


def _safe_sheet_title(title):
    cleaned = re.sub(r'[\[\]\:\*\?\/\\]', '-', str(title)).strip()
    return (cleaned or 'Turma')[:31]


def hex_to_argb(hex_color, fallback=GREEN):
    """Converte #rrggbb para AARRGGBB (openpyxl)."""
    try:
        h = str(hex_color).strip().lstrip('#')
        if len(h) == 6:
            int(h, 16)
            return 'FF' + h.upper()
    except Exception:
        pass
    return 'FF' + fallback


def _tint(hex_color, ratio=0.84):
    """Mistura a cor com branco para preservar contraste em impressão."""
    try:
        h = str(hex_color).strip().lstrip('#')
        r = int(h[0:2], 16)
        g = int(h[2:4], 16)
        b = int(h[4:6], 16)
        r = int(r + (255 - r) * ratio)
        g = int(g + (255 - g) * ratio)
        b = int(b + (255 - b) * ratio)
        return f'FF{r:02X}{g:02X}{b:02X}'
    except Exception:
        return 'FFFFFFFF'


def _border(color=LINE, style='thin'):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _build_index(aulas):
    idx = {}
    for aula in aulas:
        idx.setdefault(aula['turma_id'], {}).setdefault(aula['dia'], {})[aula['periodo']] = aula
    return idx


def _periodos_turma(turma):
    aulas_por_dia = int(turma.get('aulas_por_dia') or 5)
    return list(range(1, aulas_por_dia + 1))


def _normalize_color_mode(color_mode):
    return color_mode if color_mode in {'disciplina', 'professor', 'none'} else 'disciplina'


def _aula_color(aula, color_mode):
    if color_mode == 'professor':
        return aula.get('professor_cor')
    if color_mode == 'disciplina':
        return aula.get('disciplina_cor')
    return None


def _setup_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'B5'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.oddFooter.center.text = 'Planax - Gestão escolar'
    ws.oddFooter.right.text = 'Página &P de &N'

    for row in range(1, 16):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=PAGE)


def _write_header(ws, escola, turma):
    generated_at = datetime.now().strftime('%d/%m/%Y %H:%M')
    end_col = get_column_letter(len(_periodos_turma(turma)) + 1)
    brand_col = get_column_letter(len(_periodos_turma(turma)) + 2)

    ws.merge_cells(f'A1:{end_col}1')
    title = ws['A1']
    title.value = f'Grade de Horários - {turma["nome"]}'
    title.font = Font(name='Aptos Display', bold=True, size=20, color=INK)
    title.alignment = Alignment(horizontal='left', vertical='center')
    title.fill = PatternFill('solid', fgColor=PAGE)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{end_col}2')
    subtitle = ws['A2']
    subtitle.value = f'{escola["nome"]} | Atualizado em {generated_at}'
    subtitle.font = Font(name='Aptos', size=10, color=MUTED)
    subtitle.alignment = Alignment(horizontal='left', vertical='center')
    subtitle.fill = PatternFill('solid', fgColor=PAGE)
    ws.row_dimensions[2].height = 20

    ws[f'{brand_col}1'] = 'Planax'
    ws[f'{brand_col}1'].font = Font(name='Aptos', bold=True, size=11, color=WHITE)
    ws[f'{brand_col}1'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{brand_col}1'].fill = PatternFill('solid', fgColor=GREEN)
    ws[f'{brand_col}1'].border = _border(GREEN)

    ws[f'{brand_col}2'] = 'Gestão escolar'
    ws[f'{brand_col}2'].font = Font(name='Aptos', bold=True, size=9, color=WHITE)
    ws[f'{brand_col}2'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{brand_col}2'].fill = PatternFill('solid', fgColor=NAVY_2)
    ws[f'{brand_col}2'].border = _border(NAVY_2)


def _write_schedule(ws, turma, idx, color_mode='disciplina'):
    color_mode = _normalize_color_mode(color_mode)
    start_row = 4
    periodos = _periodos_turma(turma)
    headers = ['Dia / Período'] + [f'{periodo}º Período' for periodo in periodos]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(name='Aptos', bold=True, color=WHITE, size=10)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border(NAVY)
    ws.row_dimensions[start_row].height = 28

    ws.column_dimensions['A'].width = 18
    for col in range(2, len(periodos) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.column_dimensions[get_column_letter(len(periodos) + 2)].width = 18

    for row_idx, dia in enumerate(DIAS, start_row + 1):
        ws.row_dimensions[row_idx].height = 62

        day_cell = ws.cell(row=row_idx, column=1, value=dia)
        day_cell.font = Font(name='Aptos', bold=True, color=INK, size=10)
        day_cell.fill = PatternFill('solid', fgColor='E2E8F0')
        day_cell.alignment = Alignment(horizontal='center', vertical='center')
        day_cell.border = _border()

        for col_idx, periodo in enumerate(periodos, 2):
            aula = idx.get(turma['id'], {}).get(dia, {}).get(periodo)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _border()

            if aula:
                cor = _aula_color(aula, color_mode)
                cell.value = f"{aula['disciplina_nome']}\n{aula['professor_nome']}"
                cell.fill = PatternFill('solid', fgColor=_tint(cor) if cor else WHITE)
                cell.font = Font(
                    name='Aptos',
                    bold=True,
                    size=9,
                    color=hex_to_argb(cor, fallback=INK) if cor else INK,
                )
            else:
                cell.value = '—'
                cell.fill = PatternFill('solid', fgColor=EMPTY)
                cell.font = Font(name='Aptos', size=11, color='94A3B8')

    ws.auto_filter.ref = f'A{start_row}:{get_column_letter(len(periodos) + 1)}{start_row + len(DIAS)}'


def _write_legend(ws, turma, aulas):
    legend_title_row = 11
    ws.merge_cells(
        start_row=legend_title_row,
        start_column=1,
        end_row=legend_title_row,
        end_column=len(_periodos_turma(turma)) + 1,
    )
    title = ws.cell(row=legend_title_row, column=1, value='Legenda')
    title.font = Font(name='Aptos', bold=True, size=11, color=GREEN)
    title.alignment = Alignment(horizontal='left', vertical='center')
    title.fill = PatternFill('solid', fgColor=PAGE)

    seen = set()
    items = []
    for aula in aulas:
        if aula['turma_id'] != turma['id'] or aula['disciplina_id'] in seen:
            continue
        seen.add(aula['disciplina_id'])
        items.append(aula)

    if not items:
        cell = ws.cell(row=12, column=1, value='Nenhuma aula gerada para esta turma.')
        cell.font = Font(name='Aptos', size=9, color=MUTED)
        cell.fill = PatternFill('solid', fgColor=PAGE)
        return

    row = 12
    col = 1
    for aula in items:
        cell = ws.cell(
            row=row,
            column=col,
            value=f"{aula['disciplina_nome']}\n{aula['professor_nome']}",
        )
        cell.font = Font(name='Aptos', bold=True, size=9, color=hex_to_argb(aula.get('disciplina_cor', GREEN)))
        cell.fill = PatternFill('solid', fgColor=_tint(aula.get('disciplina_cor', GREEN), 0.9))
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = _border()
        ws.row_dimensions[row].height = 38

        col += 2
        if col > 5:
            col = 1
            row += 1


def _write_summary(wb, escola, turmas, aulas):
    ws = wb.create_sheet(title='Resumo', index=0)
    _setup_sheet(ws)

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Resumo de Horários - {escola["nome"]}'
    ws['A1'].font = Font(name='Aptos Display', bold=True, size=20, color=INK)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 32

    headers = ['Turma', 'Aulas geradas', 'Disciplinas', 'Professores']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Aptos', bold=True, size=10, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border(NAVY)

    for row_idx, turma in enumerate(turmas, 5):
        aulas_turma = [a for a in aulas if a['turma_id'] == turma['id']]
        disciplinas = sorted({a['disciplina_nome'] for a in aulas_turma})
        professores = sorted({a['professor_nome'] for a in aulas_turma})
        values = [
            turma['nome'],
            len(aulas_turma),
            ', '.join(disciplinas) or '-',
            ', '.join(professores) or '-',
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Aptos', size=9, color=INK)
            cell.fill = PatternFill('solid', fgColor=WHITE if row_idx % 2 else EMPTY)
            cell.alignment = Alignment(horizontal='left' if col_idx != 2 else 'center', vertical='center', wrap_text=True)
            cell.border = _border()
        ws.row_dimensions[row_idx].height = 34

    widths = [22, 14, 42, 42]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def exportar_excel(escola, aulas, turmas, color_mode='disciplina'):
    """Gera arquivo Excel com a grade de horários por turma."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    idx = _build_index(aulas)
    _write_summary(wb, escola, turmas, aulas)

    used_titles = set(wb.sheetnames)
    for turma in turmas:
        title = _safe_sheet_title(turma['nome'])
        original = title
        suffix = 2
        while title in used_titles:
            title = f'{original[:28]} {suffix}'[:31]
            suffix += 1
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        _setup_sheet(ws)
        _write_header(ws, escola, turma)
        _write_schedule(ws, turma, idx, color_mode=color_mode)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
